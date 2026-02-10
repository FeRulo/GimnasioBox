#!/usr/bin/env python3
"""
Script para crear el archivo Excel base del sistema de gestión del gimnasio.
Genera un archivo 'gimnasio_box.xlsx' con las pestañas, encabezados y datos de prueba.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_DATE_XLSX14
from datetime import datetime, timedelta
import hashlib
import os

# SALT debe coincidir con config.private.gs
SALT = "gimnasio_box_salt_2026_produccion_secret"

def generar_hash_password(password):
    """Genera hash SHA-256 de una contraseña + SALT"""
    data = password + SALT
    hash_obj = hashlib.sha256(data.encode('utf-8'))
    return hash_obj.hexdigest()

def crear_excel_gimnasio():
    # Crear nuevo workbook
    wb = Workbook()
    
    # Eliminar la hoja por defecto
    wb.remove(wb.active)
    
    # Definir las hojas y sus encabezados
    hojas = {
        'Clientes': [
            'Documento', 'Nombre', 'Email', 'Fecha_Nacimiento', 'Edad',
            'Contacto', 'EPS', 'Contacto_Emergencia', 'Tel_Emergencia',
            'Antecedentes', 'Objetivos', 'Objetivo_Otro', 'Estado'
        ],
        'Administradores': [
            'Documento', 'Nombre', 'Email', 'Rol', 'Password_Hash'
        ],
        'Pagos': [
            'ID_Pago', 'Fecha_Pago', 'Documento', 'Tipo_Pago',
            'Monto', 'Link_Soporte', 'Estado', 'Fecha_Aprobacion', 'Fecha_Vencimiento'
        ],
        'Horarios': [
            'ID_Clase', 'Tipo', 'Coach', 'Fecha', 'Hora', 
            'Duracion', 'Cupos_Max', 'Cupos_Reservados'
        ],
        'Reservas': [
            'ID_Reserva', 'Documento', 'ID_Clase', 
            'Fecha_Registro', 'Estado'
        ],
        'Estado_Membresias': [
            'Documento', 'Nombre', 'Tiene_Membresia_Anual', 'Vencimiento_Membresia',
            'Plan_Mensual_Activo', 'Creditos_Semanales', 'Fecha_Inicio_Plan',
            'Fecha_Vencimiento_Plan', 'Semana_Actual', 'Inicio_Semana_Actual', 'Fin_Semana_Actual',
            'Creditos_Usados_Semana', 'Creditos_Disponibles', 'Estado', 'Puede_Reservar'
        ]
    }
    
    # Estilo para los encabezados
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A2A40", end_color="1A2A40", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Crear cada hoja con sus encabezados
    for nombre_hoja, encabezados in hojas.items():
        ws = wb.create_sheet(title=nombre_hoja)
        
        # Escribir encabezados
        for col_num, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = encabezado
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    # ========== POBLAR CON DATOS DUMMY ==========
    poblar_datos_dummy(wb)
    
    # Ajustar ancho de columnas para todas las hojas
    for sheet_name in hojas.keys():
        ws = wb[sheet_name]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    # Ignorar fórmulas (empiezan con =) para calcular el ancho
                    valor = str(cell.value)
                    if not valor.startswith('='):
                        if len(valor) > max_length:
                            max_length = len(valor)
                except:
                    pass
            # Limitar ancho máximo a 25 para evitar columnas excesivamente anchas
            adjusted_width = min(max(max_length + 2, 15), 25)
            ws.column_dimensions[column].width = adjusted_width
    
    # Guardar el archivo
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    generated_dir = os.path.join(project_dir, 'generated')
    
    # Crear directorio generated si no existe
    os.makedirs(generated_dir, exist_ok=True)
    
    nombre_archivo = os.path.join(generated_dir, 'gimnasio_box.xlsx')
    wb.save(nombre_archivo)
    print(f"✅ Archivo '{nombre_archivo}' creado exitosamente!")
    print(f"📋 Hojas creadas: {', '.join(hojas.keys())}")
    print(f"📊 Datos de prueba agregados exitosamente")
    
    return nombre_archivo

def poblar_datos_dummy(wb):
    """Poblar el workbook con datos de prueba"""
    
    # ========== CLIENTES ==========
    ws_clientes = wb['Clientes']
    clientes_dummy = [
        #Documento, Nombre, Email, Fecha_Nac, Edad, Contacto, EPS, Contacto_Emerg, Tel_Emerg, Antecedentes, Objetivos, Objetivo_Otro, Estado
        ['12345678', 'Juan Pérez', 'juan.perez@mail.com', '1990-05-15', None, '3001234567', 'Salud Total', 'Pedro Pérez', '3001111111', '', 'Funcional', '', 'Activo'],
        ['87654321', 'María López', 'maria.lopez@mail.com', '1992-08-22', None, '3007654321', 'Compensar', 'Luis López', '3002222222', '', 'Perder peso', '', 'Activo'],
        ['45678912', 'Carlos Rodríguez', 'carlos.r@mail.com', '1988-03-10', None, '3009876543', 'Sura', 'Ana Rodríguez', '3003333333', '', 'Aprender boxeo', '', 'Activo'],
        ['78912345', 'Ana García', 'ana.garcia@mail.com', '1995-11-30', None, '3002345678', 'Nueva EPS', 'Jorge García', '3004444444', '', 'Otro', 'Mejorar resistencia cardiovascular', 'Activo'],
        ['11223344', 'Pedro Martínez', 'pedro.m@mail.com', '1985-07-18', None, '3003456789', 'Sanitas', 'Rosa Martínez', '3005555555', 'Hipertensión controlada', 'Funcional', '', 'Activo'],
        ['55667788', 'Laura Sánchez', 'laura.s@mail.com', '1998-12-05', None, '3004567890', 'Famisanar', 'Carlos Sánchez', '3006666666', '', 'Perder peso', '', 'Activo'],
        ['99887766', 'Diego Torres', 'diego.t@mail.com', '1993-04-25', None, '3005678901', 'Salud Total', 'Marta Torres', '3007777777', '', 'Aprender boxeo', '', 'Activo'],
        ['22334455', 'Sofia Ramírez', 'sofia.r@mail.com', '1991-09-14', None, '3006789012', 'Compensar', 'Juan Ramírez', '3008888888', '', 'Funcional', '', 'Activo'],
    ]
    
    for fila, datos in enumerate(clientes_dummy, start=2):
        for col, valor in enumerate(datos, start=1):
            cell = ws_clientes.cell(row=fila, column=col)
            # Para la columna E (Edad), insertar fórmula que calcula años desde fecha de nacimiento
            if col == 5:  # Columna E (Edad)
                # Fórmula DATEDIF para calcular edad en años
                formula = f'=DATEDIF(D{fila},TODAY(),"Y")'
                cell.value = formula
            else:
                cell.value = valor
                # Columna D (Fecha_Nacimiento): aplicar formato de fecha
                if col == 4 and valor:
                    cell.number_format = 'YYYY-MM-DD'
    
    # ========== ADMINISTRADORES ==========
    ws_admin = wb['Administradores']
    
    # NOTA IMPORTANTE: Los hashes de contraseñas se generan usando SHA-256 + SALT
    # Para agregar nuevos administradores, usa el script: scripts/generar_hash_password.py
    
    # Generar hashes para contraseñas de ejemplo
    admin_hash = generar_hash_password("admin123")  # Contraseña: admin123
    coach_hash = generar_hash_password("coach456")  # Contraseña: coach456
    
    admins_dummy = [
        # Documento, Nombre, Email, Rol, Password_Hash
        ['99999999', 'Admin Principal', 'admin@gimnasiobox.com', 'Super Admin', admin_hash],
        ['88888888', 'Hailer Joya', 'hailer@gimnasiobox.com', 'Coach', coach_hash],
    ]
    
    print(f"\n🔐 Credenciales de Administradores (GUARDAR EN LUGAR SEGURO):")
    print(f"   Usuario: 99999999 | Password: admin123 | Rol: Super Admin")
    print(f"   Usuario: 88888888 | Password: coach456 | Rol: Coach")
    print()
    
    for fila, datos in enumerate(admins_dummy, start=2):
        for col, valor in enumerate(datos, start=1):
            ws_admin.cell(row=fila, column=col, value=valor)
    
    # ========== HORARIOS ==========
    ws_horarios = wb['Horarios']
    
    # Generar horarios para los próximos 7 días
    hoy = datetime.now()
    horarios_dummy = []
    
    for dia in range(7):
        fecha = (hoy + timedelta(days=dia)).strftime('%Y-%m-%d')
        
        # Horarios matutinos
        horarios_dummy.append([
            f'CLASE{len(horarios_dummy)+1:03d}', 
            'Boxeo', 
            'Hailer', 
            fecha, 
            '07:00 AM', 
            '90 min', 
            8, 
            0
        ])
        
        horarios_dummy.append([
            f'CLASE{len(horarios_dummy)+1:03d}', 
            'Fisio', 
            'Erika', 
            fecha, 
            '09:00 AM', 
            '60 min', 
            1, 
            0
        ])
        
        # Horarios vespertinos
        horarios_dummy.append([
            f'CLASE{len(horarios_dummy)+1:03d}', 
            'Boxeo', 
            'Hailer', 
            fecha, 
            '05:00 PM', 
            '90 min', 
            8, 
            0
        ])
        
        horarios_dummy.append([
            f'CLASE{len(horarios_dummy)+1:03d}', 
            'Boxeo', 
            'Hailer', 
            fecha, 
            '06:30 PM', 
            '90 min', 
            10, 
            0
        ])
    
    for fila, datos in enumerate(horarios_dummy, start=2):
        for col, valor in enumerate(datos, start=1):
            cell = ws_horarios.cell(row=fila, column=col)
            # Para la columna H (Cupos_Reservados), insertar fórmula
            if col == 8:  # Columna H (Cupos_Reservados)
                id_clase = datos[0]
                # Fórmula que cuenta reservas activas para este ID_Clase
                formula = f'=COUNTIFS(Reservas!$C:$C,A{fila},Reservas!$E:$E,"Activa")'
                cell.value = formula
            else:
                cell.value = valor
                # Columna D (Fecha): aplicar formato de fecha
                if col == 4 and valor:
                    cell.number_format = 'YYYY-MM-DD'
    
    # ========== RESERVAS ==========
    ws_reservas = wb['Reservas']
    reservas_dummy = [
        ['RES001', '12345678', 'CLASE001', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Activa'],
        ['RES002', '87654321', 'CLASE001', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Activa'],
        ['RES003', '45678912', 'CLASE003', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Activa'],
        ['RES004', '11223344', 'CLASE005', (datetime.now() - timedelta(hours=2)).strftime('%Y-%m-%d %H:%M:%S'), 'Cancelada'],
        ['RES005', '22334455', 'CLASE002', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Activa'],
        ['RES006', '55667788', 'CLASE004', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Activa'],
        ['RES007', '99887766', 'CLASE006', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Activa'],
    ]
    
    for fila, datos in enumerate(reservas_dummy, start=2):
        for col, valor in enumerate(datos, start=1):
            cell = ws_reservas.cell(row=fila, column=col)
            cell.value = valor
            # Columna D (Fecha_Registro): aplicar formato de fecha y hora
            if col == 4 and valor:
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'
    
    # ========== ESTADO_MEMBRESIAS (con fórmulas) ==========
    ws_estado = wb['Estado_Membresias']
    
    # Obtener documentos únicos de clientes para poblar Estado_Membresias
    for fila, cliente_data in enumerate(clientes_dummy, start=2):
        documento = cliente_data[0]
        nombre = cliente_data[1]
        
        # A: Documento
        ws_estado.cell(row=fila, column=1, value=documento)
        
        # B: Nombre (VLOOKUP)
        formula_nombre = f'=VLOOKUP(A{fila},Clientes!A:B,2,FALSE)'
        ws_estado.cell(row=fila, column=2, value=formula_nombre)
        
        # C: Tiene_Membresia_Anual (SI/NO)
        formula_tiene_membresia = f'=IF(COUNTIFS(Pagos!$C:$C,A{fila},Pagos!$D:$D,"Inscripcion_Membresia_Anual",Pagos!$G:$G,"Aprobado",Pagos!$I:$I,">="&TODAY())>0,"SI","NO")'
        ws_estado.cell(row=fila, column=3, value=formula_tiene_membresia)
        
        # D: Vencimiento_Membresia (fecha de vencimiento de la membresía anual)
        formula_venc_membresia = f'=IF(COUNTIFS(Pagos!$C:$C,A{fila},Pagos!$D:$D,"Inscripcion_Membresia_Anual",Pagos!$G:$G,"Aprobado")>0,MAXIFS(Pagos!$I:$I,Pagos!$C:$C,A{fila},Pagos!$D:$D,"Inscripcion_Membresia_Anual",Pagos!$G:$G,"Aprobado"),"")'
        cell_venc_memb = ws_estado.cell(row=fila, column=4, value=formula_venc_membresia)
        cell_venc_memb.number_format = 'YYYY-MM-DD'
        
        # E: Plan_Mensual_Activo (último plan mensual aprobado)
        formula_plan = f'=IFERROR(FILTER(Pagos!D:D,Pagos!C:C=A{fila},Pagos!G:G="Aprobado",LEFT(Pagos!D:D,13)="Plan_Mensual_",Pagos!B:B=MAXIFS(Pagos!B:B,Pagos!C:C,A{fila},Pagos!G:G,"Aprobado")),"Sin Plan")'
        ws_estado.cell(row=fila, column=5, value=formula_plan)
        
        # F: Creditos_Semanales (según el plan mensual + clases individuales disponibles)
        # Suma: créditos del plan mensual + clases individuales aprobadas y vigentes
        formula_creditos = f'=IF(E{fila}="Plan_Mensual_1dia",1,IF(E{fila}="Plan_Mensual_2dias",2,IF(E{fila}="Plan_Mensual_3dias",3,IF(E{fila}="Plan_Mensual_4dias",4,IF(E{fila}="Plan_Mensual_5dias",5,0)))))+COUNTIFS(Pagos!$C:$C,A{fila},Pagos!$D:$D,"Clase_Individual",Pagos!$G:$G,"Aprobado",Pagos!$I:$I,">="&TODAY())'
        ws_estado.cell(row=fila, column=6, value=formula_creditos)
        
        # G: Fecha_Inicio_Plan (fecha del último plan mensual aprobado)
        formula_inicio = f'=IF(COUNTIFS(Pagos!$C:$C,A{fila},Pagos!$G:$G,"Aprobado")>0,MAXIFS(Pagos!$H:$H,Pagos!$C:$C,A{fila},Pagos!$G:$G,"Aprobado"),"")'
        cell_inicio = ws_estado.cell(row=fila, column=7, value=formula_inicio)
        cell_inicio.number_format = 'YYYY-MM-DD'
        
        # H: Fecha_Vencimiento_Plan (vencimiento del plan mensual actual)
        # Si tiene plan mensual activo, vence 30 días después de la fecha de inicio
        formula_venc_plan = f'=IF(E{fila}="Sin Plan","",G{fila}+30)'
        cell_venc = ws_estado.cell(row=fila, column=8, value=formula_venc_plan)
        cell_venc.number_format = 'YYYY-MM-DD'
        
        # I: Semana_Actual (semanas desde inicio del plan)
        # Usar INT() en la fecha de inicio para ignorar la hora y evitar valores negativos
        formula_semana = f'=IF(G{fila}="",0,INT((TODAY()-INT(G{fila}))/7))'
        ws_estado.cell(row=fila, column=9, value=formula_semana)
        
        # J: Inicio_Semana_Actual (columna auxiliar)
        formula_inicio_semana = f'=IF(G{fila}="","",INT(G{fila})+I{fila}*7)'
        cell_inicio_sem = ws_estado.cell(row=fila, column=10, value=formula_inicio_semana)
        cell_inicio_sem.number_format = 'YYYY-MM-DD'
        
        # K: Fin_Semana_Actual (columna auxiliar)
        formula_fin_semana = f'=IF(G{fila}="","",J{fila}+7)'
        cell_fin_sem = ws_estado.cell(row=fila, column=11, value=formula_fin_semana)
        cell_fin_sem.number_format = 'YYYY-MM-DD'
        
        # L: Creditos_Usados_Semana (fórmula simplificada usando columnas auxiliares)
        # Cuenta todas las reservas activas en la semana actual, incluyendo clases individuales
        # Usa la fecha de la clase (Horarios!D) cruzando con ID_Clase, no la fecha de registro
        formula_usados = f'=SUMPRODUCT((Reservas!$B$2:$B=A{fila})*(Reservas!$E$2:$E="Activa")*(INT(IFERROR(INDEX(Horarios!$D:$D,MATCH(Reservas!$C$2:$C,Horarios!$A:$A,0)),0))>=J{fila})*(INT(IFERROR(INDEX(Horarios!$D:$D,MATCH(Reservas!$C$2:$C,Horarios!$A:$A,0)),0))<K{fila}))'
        ws_estado.cell(row=fila, column=12, value=formula_usados)
        
        # M: Creditos_Disponibles
        formula_disponibles = f'=MAX(0,F{fila}-L{fila})'
        ws_estado.cell(row=fila, column=13, value=formula_disponibles)
        
        # N: Estado (Activo/Vencido/Sin Plan)
        # Considera activo si tiene plan mensual válido O clases individuales disponibles
        formula_estado = f'=IF(AND(E{fila}="Sin Plan",F{fila}=0),"Sin Plan",IF(AND(E{fila}<>"Sin Plan",TODAY()>H{fila}),"Vencido",IF(AND(C{fila}="NO",E{fila}<>"Sin Plan"),"Sin Membresia","Activo")))'
        ws_estado.cell(row=fila, column=14, value=formula_estado)
        
        # O: Puede_Reservar
        formula_puede = f'=IF(AND(N{fila}="Activo",M{fila}>0),"SI","NO")'
        ws_estado.cell(row=fila, column=15, value=formula_puede)
    
    # ========== PAGOS ==========
    ws_pagos = wb['Pagos']
    fecha_hoy = datetime.now()
    pagos_dummy = [
        # ID_Pago, Fecha_Pago, Documento, Tipo_Pago, Monto, Link_Soporte, Estado, Fecha_Aprobacion, Fecha_Vencimiento
        # Primero pagan la membresía anual, luego los planes mensuales
        ['PAG001', fecha_hoy - timedelta(days=30), '12345678', 'Inscripcion_Membresia_Anual', 35000, 'https://drive.google.com/file/ejemplo1', 'Aprobado', fecha_hoy - timedelta(days=30), None],
        ['PAG002', fecha_hoy, '12345678', 'Plan_Mensual_2dias', 145000, 'https://drive.google.com/file/ejemplo2', 'Aprobado', fecha_hoy, None],
        ['PAG003', fecha_hoy - timedelta(days=10), '87654321', 'Inscripcion_Membresia_Anual', 35000, 'https://drive.google.com/file/ejemplo3', 'Aprobado', fecha_hoy - timedelta(days=10), None],
        ['PAG004', fecha_hoy - timedelta(days=9), '87654321', 'Plan_Mensual_3dias', 175000, 'https://drive.google.com/file/ejemplo4', 'Aprobado', fecha_hoy - timedelta(days=9), None],
        ['PAG005', fecha_hoy - timedelta(days=1), '45678912', 'Clase_Individual', 45000, 'https://drive.google.com/file/ejemplo5', 'Aprobado', fecha_hoy - timedelta(days=1), None],
        ['PAG006', fecha_hoy, '78912345', 'Plan_Mensual_1dia', 115000, 'https://drive.google.com/file/ejemplo6', 'Pendiente', '', ''],
    ]
    
    for fila, datos in enumerate(pagos_dummy, start=2):
        for col, valor in enumerate(datos, start=1):
            cell = ws_pagos.cell(row=fila, column=col)
            
            # Para la columna I (Fecha_Vencimiento), insertar fórmula
            if col == 9 and datos[6] == 'Aprobado':  # Solo si está aprobado
                tipo_pago = datos[3]
                if 'Membresia_Anual' in tipo_pago or 'Inscripcion' in tipo_pago:
                    # Membresía anual vence a 365 días
                    formula = f'=H{fila}+365'
                elif 'Plan_Mensual' in tipo_pago:
                    # Planes mensuales vencen a 30 días
                    formula = f'=H{fila}+30'
                elif 'Clase_Individual' in tipo_pago:
                    # Clases individuales vencen a 30 días
                    formula = f'=H{fila}+30'
                else:
                    formula = ''
                cell.value = formula
                if formula:  # Aplicar formato de fecha si hay fórmula
                    cell.number_format = 'YYYY-MM-DD'
            else:
                cell.value = valor
                # Aplicar formato de fecha a columnas B (Fecha_Pago), H (Fecha_Aprobacion) e I (Fecha_Vencimiento)
                if col == 2 and isinstance(valor, datetime):  # Fecha_Pago
                    cell.number_format = 'YYYY-MM-DD'
                elif col == 8 and isinstance(valor, datetime):  # Fecha_Aprobacion
                    cell.number_format = 'YYYY-MM-DD'
                elif col == 9:  # Fecha_Vencimiento (incluso si está vacía, preparar formato)
                    cell.number_format = 'YYYY-MM-DD'

if __name__ == "__main__":
    crear_excel_gimnasio()
